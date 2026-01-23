from flask import Flask, render_template, request, session, redirect, url_for, flash, blueprint
import sqlite3
import hashlib
import json
from datetime import datetime
from openpyxl import load_workbook

app = Flask(__name__)
app.secret_key = "super-secret-key-change-me"

EXCEL_FILE = "plswork.xlsx"
DB_FILE = "app.db"

# -------------------- DATABASE --------------------

def get_db_connection():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    return conn

# -------------------- EXCEL HELPERS --------------------

def read_excel_sheet(sheet_name):
    try:
        wb = load_workbook(EXCEL_FILE, data_only=True)
        sheet = wb[sheet_name]

        headers = [cell.value for cell in sheet[1]]
        rows = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, row)))

        return rows
    except Exception as e:
        print(f"Excel error ({sheet_name}):", e)
        return []

def safe_date(val):
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return str(val)

# -------------------- AUTH --------------------

@app.route("/")
def login_page():
    return render_template("login.html")

@app.route("/login", methods=["POST"])
def login():
    username = request.form["username"]
    password = request.form["password"]
    password_hash = hashlib.sha256(password.encode()).hexdigest()

    conn = get_db_connection()
    user = conn.execute(
        "SELECT * FROM users WHERE username=? AND password_hash=?",
        (username, password_hash)
    ).fetchone()
    conn.close()

    if user:
        session["username"] = username
        return redirect(url_for("dashboard"))

    # Excel fallback
    users = read_excel_sheet("users")
    for u in users:
        if u.get("username") == username and u.get("password") == password:
            session["username"] = username
            return redirect(url_for("dashboard"))

    flash("Invalid username or password")
    return redirect(url_for("login_page"))

# -------------------- LightMode / Darkmode --------------------
@pages.get("/toggle-theme")
def toggle_theme():
    current_theme = session.get("theme")
    if current_theme == "dark":
        session["theme"] = "light"
    else:
        session["theme"] = "dark"

    return redirect(request.args.get("current_page"))
# -------------------- DASHBOARD --------------------

@app.route("/dashboard")
def dashboard():
    if "username" not in session:
        return redirect(url_for("login_page"))

    # KPI
    kpi_rows = read_excel_sheet("data")
    kpi_labels = [safe_date(r.get("date")) for r in kpi_rows]
    kpi_values = [float(r.get("kpi_rate", 0)) for r in kpi_rows]

    change = kpi_values[-1] - kpi_values[-2] if len(kpi_values) >= 2 else 0

    # Sales
    sales_rows = read_excel_sheet("sales")
    sales_map = {}
    for r in sales_rows:
        product = r.get("product")
        units = int(r.get("units_sold") or 0)
        sales_map[product] = sales_map.get(product, 0) + units

    sales_products = list(sales_map.keys())
    sales_units = list(sales_map.values())

    # Employees
    employee_rows = read_excel_sheet("employee")
    employee_rows.sort(
        key=lambda x: float(str(x.get("efficiency", 0)).replace("%", "") or 0),
        reverse=True
    )
    top = employee_rows[:5]
    employee_names = [e.get("name") for e in top]
    employee_efficiency = [
        float(str(e.get("efficiency", 0)).replace("%", "") or 0)
        for e in top
    ]

    # Tasks status
    task_rows = read_excel_sheet("tasks")
    status_count = {}
    for r in task_rows:
        s = r.get("status", "Unknown")
        status_count[s] = status_count.get(s, 0) + 1

    task_labels = list(status_count.keys())
    task_counts = list(status_count.values())

    return render_template(
        "index.html",
        username=session["username"],
        kpi_labels=json.dumps(kpi_labels),
        kpi_values=json.dumps(kpi_values),
        change=change,
        sales_products=json.dumps(sales_products),
        sales_units=json.dumps(sales_units),
        employee_names=json.dumps(employee_names),
        employee_efficiency=json.dumps(employee_efficiency),
        task_labels=json.dumps(task_labels),
        task_counts=json.dumps(task_counts),
    )

# -------------------- TASKS --------------------

@app.route("/tasks")
def tasks():
    if "username" not in session:
        return redirect(url_for("login_page"))

    username = session["username"]
    conn = get_db_connection()
    tasks = conn.execute(
        "SELECT * FROM tasks WHERE assigned_to=?",
        (username,)
    ).fetchall()
    conn.close()

    task_list = [dict(t) for t in tasks]

    if not task_list:
        excel_tasks = read_excel_sheet("tasks")
        for i, t in enumerate(excel_tasks):
            if username in str(t.get("assigned_to", "")):
                task_list.append({
                    "id": i,
                    "task": t.get("task"),
                    "status": t.get("status") or "Not Started"
                })

    return render_template("tasks.html", username=username, tasks=task_list)

@app.route("/add_task", methods=["POST"])
def add_task():
    if "username" not in session:
        return redirect(url_for("login_page"))

    conn = get_db_connection()
    conn.execute(
        "INSERT INTO tasks (task, status, assigned_to) VALUES (?, ?, ?)",
        (request.form["task"], request.form.get("status", "Not Started"), session["username"])
    )
    conn.commit()
    conn.close()
    return redirect(url_for("tasks"))

@app.route("/update_task/<int:task_id>", methods=["POST"])
def update_task(task_id):
    conn = get_db_connection()
    conn.execute(
        "UPDATE tasks SET status=? WHERE id=? AND assigned_to=?",
        (request.form["status"], task_id, session["username"])
    )
    conn.commit()
    conn.close()
    return redirect(url_for("tasks"))

@app.route("/delete_task/<int:task_id>", methods=["POST"])
def delete_task(task_id):
    conn = get_db_connection()
    conn.execute(
        "DELETE FROM tasks WHERE id=? AND assigned_to=?",
        (task_id, session["username"])
    )
    conn.commit()
    conn.close()
    return redirect(url_for("tasks"))

# -------------------- MESSAGES --------------------

@app.route("/messages")
def messages():
    if "username" not in session:
        return redirect(url_for("login_page"))

    username = session["username"]
    conn = get_db_connection()

    conversations = conn.execute("""
        SELECT DISTINCT
        CASE WHEN sender=? THEN recipient ELSE sender END AS other_user
        FROM messages
        WHERE sender=? OR recipient=?
    """, (username, username, username)).fetchall()

    messages = []
    active = None

    if conversations:
        active = conversations[0]["other_user"]
        messages = conn.execute("""
            SELECT * FROM messages
            WHERE (sender=? AND recipient=?) OR (sender=? AND recipient=?)
            ORDER BY timestamp
        """, (username, active, active, username)).fetchall()

    conn.close()

    return render_template(
        "Messages.html",
        username=username,
        conversations=conversations,
        messages=messages,
        active_conversation=active
    )

# -------------------- SETTINGS / LOGOUT --------------------

@app.route("/settings")
def settings():
    if "username" not in session:
        return redirect(url_for("login_page"))
    return render_template("Settings.html", username=session["username"])

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login_page"))

# -------------------- RUN --------------------

if __name__ == "__main__":
    app.run()


